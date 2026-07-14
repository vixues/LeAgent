"""Excel Generator Tool - Create Excel files programmatically.

Uses openpyxl for creating .xlsx files with multiple sheets, formatting, and charts.
"""

from __future__ import annotations

import subprocess
from pathlib import Path
from typing import Any

import structlog

from leagent.tools.base import SyncTool, ToolCategory, ToolContext

logger = structlog.get_logger(__name__)

_FINANCIAL_NUMBER_FORMATS: dict[str, str] = {
    "currency": '$#,##0;($#,##0);"-"',
    "currency_cents": '$#,##0.00;($#,##0.00);"-"',
    "percentage": "0.0%",
    "multiple": "0.0x",
    "integer": '#,##0;(#,##0);"-"',
    "year": "@",
}


def _apply_financial_preset(
    ws: Any,
    headers: list[str],
    data: list[list[Any]],
    Font: Any,
    PatternFill: Any,
    Alignment: Any,
) -> None:
    """Apply IB-standard color coding and formatting to a financial sheet."""
    blue_font = Font(color="0000FF", name="Arial", size=11)
    black_font = Font(color="000000", name="Arial", size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, name="Arial", size=11)
    header_align = Alignment(horizontal="center")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    start_row = 2 if headers else 1
    for row_idx, row_data in enumerate(data, start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            is_formula = isinstance(value, str) and str(value).startswith("=")
            cell.font = black_font if is_formula else blue_font


def _try_recalculate(output_path: Path) -> dict[str, Any] | None:
    """Attempt formula recalculation via scripts/recalc.py (requires LibreOffice)."""
    recalc_candidates = [
        Path(__file__).resolve().parents[3] / "scripts" / "recalc.py",
        Path.cwd() / "scripts" / "recalc.py",
    ]
    recalc_script: Path | None = None
    for cand in recalc_candidates:
        if cand.is_file():
            recalc_script = cand
            break

    if recalc_script is None:
        logger.debug("recalc_script_not_found")
        return None

    try:
        result = subprocess.run(
            ["python", str(recalc_script), str(output_path), "30"],
            capture_output=True,
            text=True,
            timeout=60,
        )
        if result.returncode == 0:
            import json

            try:
                return json.loads(result.stdout.strip())
            except Exception:
                return {"status": "unknown", "stdout": result.stdout[:500]}
        logger.warning(
            "recalc_failed",
            returncode=result.returncode,
            stderr=result.stderr[:500],
        )
    except FileNotFoundError:
        logger.debug("recalc_python_not_found")
    except subprocess.TimeoutExpired:
        logger.warning("recalc_timeout")
    except Exception:
        logger.exception("recalc_unexpected_error")
    return None


class ExcelGeneratorTool(SyncTool):
    """Generate Excel files (.xlsx) with rich formatting and charts.

    Features:
    - Create workbooks with multiple sheets
    - Cell formatting (fonts, colors, borders, alignment)
    - Formulas and calculated cells
    - Basic chart generation (bar, line, pie)
    - Conditional formatting
    - Freeze panes and filters
    - Merge cells and row heights
    - Data validation (dropdowns, numeric ranges)
    - Financial model preset with IB-standard color coding
    - Optional formula recalculation via LibreOffice
    """

    name = "excel_generator"
    description = (
        "Generate Excel files (.xlsx) with multiple sheets, cell formatting, "
        "formulas, charts, and conditional formatting. Use formulas instead of "
        "hardcoded values so the spreadsheet stays dynamic. Supports merged "
        "cells, data validation, row heights, and a financial model preset "
        "with industry-standard color coding (blue=inputs, black=formulas). "
        "Optionally recalculates formulas via LibreOffice."
    )
    category = ToolCategory.GEN
    version = "2.0.0"
    timeout_sec = 180
    aliases = ["xlsx_gen", "spreadsheet_gen", "create_excel"]
    search_hint = (
        "Excel xlsx spreadsheet generate create formatting charts formulas "
        "sheets financial model budget forecast"
    )
    is_concurrency_safe = False
    is_read_only = False
    interrupt_behavior = "cancel"
    max_result_size_chars = 50_000
    output_path_params = ("output_path",)

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "output_path": {
                    "type": "string",
                    "description": (
                        "Use a bare filename for the generated Excel file "
                        "(for example, 'workbook.xlsx'); it will be placed in the "
                        "session workspace and shown in the Files tab. Use only "
                        "when the user asked to save or export."
                    ),
                },
                "preset": {
                    "type": "string",
                    "enum": ["financial"],
                    "description": (
                        "Apply a preset. 'financial' sets IB-standard color coding "
                        "(blue text = hardcoded inputs, black = formulas), Arial font, "
                        "and financial number formats."
                    ),
                },
                "recalculate_formulas": {
                    "type": "boolean",
                    "description": (
                        "After saving, recalculate all formulas via LibreOffice. "
                        "Requires LibreOffice installed. Returns error summary if "
                        "formula errors are found."
                    ),
                },
                "sheets": {
                    "type": "array",
                    "description": "Array of sheet definitions.",
                    "items": {
                        "type": "object",
                        "properties": {
                            "name": {
                                "type": "string",
                                "description": "Sheet name (max 31 characters).",
                            },
                            "data": {
                                "type": "array",
                                "description": "2D array of cell values (rows x columns).",
                                "items": {
                                    "type": "array",
                                    "items": {},
                                },
                            },
                            "headers": {
                                "type": "array",
                                "description": "Header row (will be styled as headers).",
                                "items": {"type": "string"},
                            },
                            "column_widths": {
                                "type": "object",
                                "description": "Column widths as {column_letter: width}.",
                                "additionalProperties": {"type": "number"},
                            },
                            "row_heights": {
                                "type": "object",
                                "description": "Row heights as {row_number_string: height_in_points}.",
                                "additionalProperties": {"type": "number"},
                            },
                            "freeze_panes": {
                                "type": "string",
                                "description": "Cell reference for freeze panes (e.g., 'A2' freezes first row).",
                            },
                            "auto_filter": {
                                "type": "boolean",
                                "description": "Enable auto-filter on data range.",
                            },
                            "formulas": {
                                "type": "array",
                                "description": "Array of formula definitions.",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "cell": {
                                            "type": "string",
                                            "description": "Cell reference (e.g., 'E2').",
                                        },
                                        "formula": {
                                            "type": "string",
                                            "description": "Excel formula (e.g., '=SUM(A2:D2)').",
                                        },
                                    },
                                    "required": ["cell", "formula"],
                                },
                            },
                            "merged_cells": {
                                "type": "array",
                                "description": "Ranges to merge (e.g., ['A1:C1', 'D1:F1']).",
                                "items": {"type": "string"},
                            },
                            "data_validation": {
                                "type": "array",
                                "description": "Data validation rules.",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "range": {
                                            "type": "string",
                                            "description": "Cell range (e.g., 'B2:B100').",
                                        },
                                        "type": {
                                            "type": "string",
                                            "enum": ["list", "whole", "decimal"],
                                            "description": "Validation type.",
                                        },
                                        "formula1": {
                                            "type": "string",
                                            "description": "Comma-separated list items or min value.",
                                        },
                                        "formula2": {
                                            "type": "string",
                                            "description": "Max value (for whole/decimal).",
                                        },
                                        "operator": {
                                            "type": "string",
                                            "enum": ["between", "greaterThan", "lessThan", "equal"],
                                        },
                                    },
                                    "required": ["range", "type"],
                                },
                            },
                            "cell_styles": {
                                "type": "array",
                                "description": "Array of cell style definitions.",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "range": {
                                            "type": "string",
                                            "description": "Cell range (e.g., 'A1:D1' or 'A1').",
                                        },
                                        "bold": {"type": "boolean"},
                                        "italic": {"type": "boolean"},
                                        "font_size": {"type": "integer"},
                                        "font_name": {"type": "string"},
                                        "font_color": {
                                            "type": "string",
                                            "description": "Hex color code (e.g., 'FF0000').",
                                        },
                                        "bg_color": {
                                            "type": "string",
                                            "description": "Background hex color (e.g., 'FFFF00').",
                                        },
                                        "alignment": {
                                            "type": "string",
                                            "enum": ["left", "center", "right"],
                                        },
                                        "number_format": {
                                            "type": "string",
                                            "description": (
                                                "Excel number format (e.g., '#,##0.00', '0%'). "
                                                "Financial shortcuts: 'currency', 'currency_cents', "
                                                "'percentage', 'multiple', 'integer', 'year'."
                                            ),
                                        },
                                        "border": {
                                            "type": "string",
                                            "enum": ["thin", "medium", "thick"],
                                        },
                                    },
                                    "required": ["range"],
                                },
                            },
                            "charts": {
                                "type": "array",
                                "description": "Array of chart definitions.",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "type": {
                                            "type": "string",
                                            "enum": ["bar", "line", "pie", "column", "area"],
                                            "description": "Chart type.",
                                        },
                                        "title": {
                                            "type": "string",
                                            "description": "Chart title.",
                                        },
                                        "data_range": {
                                            "type": "string",
                                            "description": "Data range for chart (e.g., 'A1:B10').",
                                        },
                                        "categories_range": {
                                            "type": "string",
                                            "description": "Range for category labels.",
                                        },
                                        "position": {
                                            "type": "string",
                                            "description": "Cell where chart top-left is anchored (e.g., 'E2').",
                                        },
                                        "width": {
                                            "type": "integer",
                                            "description": "Chart width in cells.",
                                        },
                                        "height": {
                                            "type": "integer",
                                            "description": "Chart height in cells.",
                                        },
                                    },
                                    "required": ["type", "data_range", "position"],
                                },
                            },
                            "conditional_formatting": {
                                "type": "array",
                                "description": "Conditional formatting rules.",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "range": {
                                            "type": "string",
                                            "description": "Cell range to apply formatting.",
                                        },
                                        "rule_type": {
                                            "type": "string",
                                            "enum": ["cell_value", "color_scale", "data_bar"],
                                        },
                                        "operator": {
                                            "type": "string",
                                            "enum": ["greaterThan", "lessThan", "equal", "between"],
                                        },
                                        "value": {
                                            "type": ["string", "number"],
                                            "description": "Comparison value.",
                                        },
                                        "format_color": {
                                            "type": "string",
                                            "description": "Fill color when condition is met.",
                                        },
                                    },
                                    "required": ["range", "rule_type"],
                                },
                            },
                        },
                        "required": ["name"],
                    },
                },
                "workbook_properties": {
                    "type": "object",
                    "description": "Workbook-level properties.",
                    "properties": {
                        "title": {"type": "string"},
                        "author": {"type": "string"},
                        "subject": {"type": "string"},
                        "company": {"type": "string"},
                    },
                },
            },
            "required": ["output_path", "sheets"],
            "additionalProperties": False,
        }

    def get_activity_description(self, params: dict[str, Any] | None = None) -> str | None:
        path = (params or {}).get("output_path", "")
        return f"Generating Excel file{f': {path}' if path else ''}"

    def execute_sync(self, params: dict[str, Any], context: ToolContext) -> dict[str, Any]:
        """Generate an Excel file with the specified sheets and formatting."""
        try:
            from openpyxl import Workbook
            from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart
            from openpyxl.chart.reference import Reference
            from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.datavalidation import DataValidation
        except ImportError as e:
            raise RuntimeError(
                "openpyxl is not installed. Install with: pip install openpyxl"
            ) from e

        output_path = Path(params["output_path"])
        sheets_config = params.get("sheets", [])
        wb_props = params.get("workbook_properties", {})
        preset = params.get("preset")
        recalculate = params.get("recalculate_formulas", False)

        output_path.parent.mkdir(parents=True, exist_ok=True)

        logger.info("Creating Excel workbook", output_path=str(output_path))

        wb = Workbook()
        default_sheet = wb.active

        default_font_name = "Arial" if preset == "financial" else "Calibri"
        wb.default_font = Font(name=default_font_name, size=11)

        if wb_props:
            if wb_props.get("title"):
                wb.properties.title = wb_props["title"]
            if wb_props.get("author"):
                wb.properties.creator = wb_props["author"]
            if wb_props.get("subject"):
                wb.properties.subject = wb_props["subject"]
            if wb_props.get("company"):
                wb.properties.company = wb_props["company"]

        stats = {
            "sheets": 0,
            "total_rows": 0,
            "total_cells": 0,
            "charts": 0,
            "formulas": 0,
            "merged_ranges": 0,
            "validations": 0,
        }

        chart_classes = {
            "bar": BarChart,
            "column": BarChart,
            "line": LineChart,
            "pie": PieChart,
            "area": AreaChart,
        }

        for idx, sheet_config in enumerate(sheets_config):
            sheet_name = sheet_config.get("name", f"Sheet{idx + 1}")[:31]

            if idx == 0 and default_sheet:
                ws = default_sheet
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)

            headers = sheet_config.get("headers", [])
            data = sheet_config.get("data", [])

            if headers:
                header_font = Font(bold=True, name=default_font_name)
                header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
                header_align = Alignment(horizontal="center")
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    stats["total_cells"] += 1
                stats["total_rows"] += 1

            start_row = 2 if headers else 1
            for row_idx, row_data in enumerate(data, start_row):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    stats["total_cells"] += 1
                stats["total_rows"] += 1

            if preset == "financial" and (headers or data):
                _apply_financial_preset(ws, headers, data, Font, PatternFill, Alignment)

            for col_letter, width in sheet_config.get("column_widths", {}).items():
                ws.column_dimensions[col_letter].width = width

            for row_str, height in sheet_config.get("row_heights", {}).items():
                try:
                    ws.row_dimensions[int(row_str)].height = height
                except (ValueError, TypeError):
                    logger.warning("invalid_row_height_key", key=row_str)

            if sheet_config.get("freeze_panes"):
                ws.freeze_panes = sheet_config["freeze_panes"]

            if sheet_config.get("auto_filter") and (headers or data):
                max_row = len(data) + (1 if headers else 0)
                max_col = len(headers) if headers else (len(data[0]) if data else 0)
                if max_col > 0:
                    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

            for formula_def in sheet_config.get("formulas", []):
                cell_ref = formula_def["cell"]
                formula = formula_def["formula"]
                ws[cell_ref] = formula
                stats["formulas"] += 1

            for merge_range in sheet_config.get("merged_cells", []):
                ws.merge_cells(merge_range)
                stats["merged_ranges"] += 1

            for dv_def in sheet_config.get("data_validation", []):
                dv = self._create_data_validation(dv_def, DataValidation)
                if dv:
                    ws.add_data_validation(dv)
                    dv.add(dv_def["range"])
                    stats["validations"] += 1

            for style_def in sheet_config.get("cell_styles", []):
                self._apply_cell_style(ws, style_def, Font, PatternFill, Alignment, Border, Side)

            for chart_def in sheet_config.get("charts", []):
                chart_type = chart_def["type"]
                chart_class = chart_classes.get(chart_type)

                if not chart_class:
                    logger.warning("Unknown chart type", chart_type=chart_type)
                    continue

                chart = chart_class()
                if chart_def.get("title"):
                    chart.title = chart_def["title"]

                data_range = chart_def["data_range"]
                min_col, min_row, max_col, max_row = self._parse_range(data_range)

                data_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)

                chart.add_data(data_ref, titles_from_data=True)
                if chart_def.get("categories_range"):
                    cat_min_col, cat_min_row, cat_max_col, cat_max_row = self._parse_range(
                        chart_def["categories_range"]
                    )
                    cats = Reference(
                        ws,
                        min_col=cat_min_col,
                        min_row=cat_min_row,
                        max_col=cat_max_col,
                        max_row=cat_max_row,
                    )
                    chart.set_categories(cats)

                if chart_type == "column":
                    chart.type = "col"

                if chart_def.get("width"):
                    chart.width = chart_def["width"]
                if chart_def.get("height"):
                    chart.height = chart_def["height"]

                ws.add_chart(chart, chart_def["position"])
                stats["charts"] += 1

            for cf_def in sheet_config.get("conditional_formatting", []):
                self._apply_conditional_formatting(
                    ws, cf_def, CellIsRule, ColorScaleRule, DataBarRule, PatternFill
                )

            stats["sheets"] += 1

        wb.save(str(output_path))

        recalc_result: dict[str, Any] | None = None
        if recalculate:
            recalc_result = _try_recalculate(output_path)

        file_size = output_path.stat().st_size

        logger.info(
            "Excel file generated successfully",
            output_path=str(output_path),
            file_size=file_size,
            preset=preset or "none",
            **stats,
        )

        result: dict[str, Any] = {
            "success": True,
            "output_path": str(output_path),
            "file_size_bytes": file_size,
            "stats": stats,
            "sheet_names": [s.get("name", f"Sheet{i+1}")[:31] for i, s in enumerate(sheets_config)],
        }
        if recalc_result is not None:
            result["recalc"] = recalc_result
        # Quality / managed identity are owned by SessionManager after promotion
        # (ArtifactRegistrar → register_external_file); do not gate here.
        return result

    def _parse_range(self, range_str: str) -> tuple[int, int, int, int]:
        """Parse Excel range string into column/row bounds."""
        from openpyxl.utils import range_boundaries

        return range_boundaries(range_str)

    def _create_data_validation(
        self, dv_def: dict[str, Any], DataValidation: Any
    ) -> Any | None:
        """Create an openpyxl DataValidation from a definition dict."""
        dv_type = dv_def.get("type", "list")
        if dv_type == "list":
            formula1 = dv_def.get("formula1", "")
            if formula1 and not formula1.startswith('"'):
                formula1 = f'"{formula1}"'
            return DataValidation(type="list", formula1=formula1, allow_blank=True)
        elif dv_type in ("whole", "decimal"):
            operator = dv_def.get("operator", "between")
            f1 = dv_def.get("formula1")
            f2 = dv_def.get("formula2")
            return DataValidation(
                type=dv_type, operator=operator, formula1=f1, formula2=f2, allow_blank=True
            )
        return None

    def _apply_cell_style(
        self,
        ws: Any,
        style_def: dict[str, Any],
        Font: Any,
        PatternFill: Any,
        Alignment: Any,
        Border: Any,
        Side: Any,
    ) -> None:
        """Apply style to a cell range."""
        from openpyxl.utils import range_boundaries

        range_str = style_def["range"]

        if ":" in range_str:
            min_col, min_row, max_col, max_row = range_boundaries(range_str)
        else:
            min_col, min_row, max_col, max_row = range_boundaries(f"{range_str}:{range_str}")

        font_kwargs: dict[str, Any] = {}
        if style_def.get("bold"):
            font_kwargs["bold"] = True
        if style_def.get("italic"):
            font_kwargs["italic"] = True
        if style_def.get("font_size"):
            font_kwargs["size"] = style_def["font_size"]
        if style_def.get("font_name"):
            font_kwargs["name"] = style_def["font_name"]
        if style_def.get("font_color"):
            font_kwargs["color"] = style_def["font_color"]

        fill = None
        if style_def.get("bg_color"):
            fill = PatternFill(
                start_color=style_def["bg_color"],
                end_color=style_def["bg_color"],
                fill_type="solid",
            )

        alignment = None
        if style_def.get("alignment"):
            alignment = Alignment(horizontal=style_def["alignment"])

        border = None
        if style_def.get("border"):
            side = Side(style=style_def["border"])
            border = Border(left=side, right=side, top=side, bottom=side)

        raw_fmt = style_def.get("number_format")
        number_format = _FINANCIAL_NUMBER_FORMATS.get(raw_fmt, raw_fmt) if raw_fmt else None

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if font_kwargs:
                    cell.font = Font(**font_kwargs)
                if fill:
                    cell.fill = fill
                if alignment:
                    cell.alignment = alignment
                if border:
                    cell.border = border
                if number_format:
                    cell.number_format = number_format

    def _apply_conditional_formatting(
        self,
        ws: Any,
        cf_def: dict[str, Any],
        CellIsRule: Any,
        ColorScaleRule: Any,
        DataBarRule: Any,
        PatternFill: Any,
    ) -> None:
        """Apply conditional formatting rule."""
        range_str = cf_def["range"]
        rule_type = cf_def["rule_type"]

        if rule_type == "cell_value":
            operator = cf_def.get("operator", "greaterThan")
            value = cf_def.get("value", 0)
            color = cf_def.get("format_color", "FFFF00")

            rule = CellIsRule(
                operator=operator,
                formula=[str(value)],
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
            )
            ws.conditional_formatting.add(range_str, rule)

        elif rule_type == "color_scale":
            rule = ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            )
            ws.conditional_formatting.add(range_str, rule)

        elif rule_type == "data_bar":
            rule = DataBarRule(
                start_type="min",
                end_type="max",
                color="638EC6",
            )
            ws.conditional_formatting.add(range_str, rule)
