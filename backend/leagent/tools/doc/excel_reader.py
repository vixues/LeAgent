"""Excel Reader Tool - Read and extract data from Excel files.

Uses openpyxl for .xlsx file processing with sheet selection and range extraction.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import structlog

from leagent.tools.base import SyncTool, ToolCategory, ToolContext, ValidationResult

logger = structlog.get_logger(__name__)


class ExcelReaderTool(SyncTool):
    """Read and extract data from Excel files (.xlsx).

    Features:
    - Sheet selection by name or index
    - Cell range extraction
    - Return as dict or DataFrame-compatible format
    - Header row detection
    - Multiple sheet reading
    """

    name = "excel_reader"
    description = (
        "Read data from Excel files (.xlsx) with support for sheet selection, "
        "cell range extraction, and flexible output formats."
    )
    category = ToolCategory.DOC
    version = "1.0.0"
    timeout_sec = 120
    aliases = ["xlsx_reader", "spreadsheet_reader", "read_excel"]
    search_hint = "Excel xlsx spreadsheet read sheet cell range data extract"
    is_concurrency_safe = True
    is_read_only = True
    interrupt_behavior = "cancel"
    max_result_size_chars = 200_000
    path_params = ("file_path",)

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path to the Excel file (.xlsx) to read.",
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Name of the sheet to read. If not specified, reads the active sheet.",
                },
                "sheet_index": {
                    "type": "integer",
                    "description": "Index of the sheet to read (0-indexed). Ignored if sheet_name is provided.",
                    "minimum": 0,
                },
                "cell_range": {
                    "type": "string",
                    "description": "Cell range to extract (e.g., 'A1:D10'). If not specified, reads all data.",
                },
                "has_header": {
                    "type": "boolean",
                    "description": "Whether the first row contains headers. Defaults to True.",
                    "default": True,
                },
                "output_format": {
                    "type": "string",
                    "enum": ["dict", "records", "list"],
                    "description": (
                        "Output format: 'dict' (column-oriented), 'records' (row-oriented list of dicts), "
                        "'list' (raw 2D list). Defaults to 'records'."
                    ),
                    "default": "records",
                },
                "include_empty_cells": {
                    "type": "boolean",
                    "description": "Whether to include empty cells as None. Defaults to True.",
                    "default": True,
                },
            },
            "required": ["file_path"],
            "additionalProperties": False,
        }

    def get_activity_description(self, params: dict[str, Any] | None = None) -> str | None:
        return "Reading Excel file"

    async def validate_input(self, params: dict[str, Any], context: ToolContext) -> ValidationResult:
        from pathlib import Path
        fp = params.get("file_path", "")
        if fp and not Path(fp).exists():
            return ValidationResult(valid=False, message=f"File not found: {fp}")
        return ValidationResult(valid=True)

    def execute_sync(self, params: dict[str, Any], context: ToolContext) -> dict[str, Any]:
        """Read data from an Excel file.

        Args:
            params: Tool parameters including file_path and reading options.
            context: Execution context.

        Returns:
            Dictionary containing extracted data and workbook information.

        Raises:
            FileNotFoundError: If the Excel file doesn't exist.
            ValueError: If the file format is invalid or sheet not found.
            RuntimeError: If openpyxl encounters an error.
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import range_boundaries
            from openpyxl.utils.exceptions import InvalidFileException
        except ImportError as e:
            raise RuntimeError(
                "openpyxl is not installed. Install with: pip install openpyxl"
            ) from e

        file_path = Path(params["file_path"])
        sheet_name = params.get("sheet_name")
        sheet_index = params.get("sheet_index", 0)
        cell_range = params.get("cell_range")
        has_header = params.get("has_header", True)
        output_format = params.get("output_format", "records")
        include_empty_cells = params.get("include_empty_cells", True)

        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        if file_path.suffix.lower() not in (".xlsx", ".xlsm"):
            raise ValueError(
                f"Unsupported file format: {file_path.suffix}. "
                "Only .xlsx and .xlsm files are supported. For .xls files, convert to .xlsx first."
            )

        logger.info("Opening Excel file", file_path=str(file_path))

        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
        except InvalidFileException as e:
            raise ValueError(f"Invalid or corrupted Excel file: {file_path}") from e
        except Exception as e:
            raise RuntimeError(f"Failed to open Excel file: {e}") from e

        try:
            sheet_names = wb.sheetnames

            if sheet_name:
                if sheet_name not in sheet_names:
                    raise ValueError(
                        f"Sheet '{sheet_name}' not found. Available sheets: {sheet_names}"
                    )
                ws = wb[sheet_name]
                actual_sheet_name = sheet_name
            elif sheet_index < len(sheet_names):
                actual_sheet_name = sheet_names[sheet_index]
                ws = wb[actual_sheet_name]
            else:
                raise ValueError(
                    f"Sheet index {sheet_index} out of range. "
                    f"Workbook has {len(sheet_names)} sheet(s)."
                )

            if cell_range:
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
                except Exception as e:
                    raise ValueError(f"Invalid cell range '{cell_range}': {e}") from e
                rows = ws.iter_rows(
                    min_row=min_row, max_row=max_row,
                    min_col=min_col, max_col=max_col
                )
            else:
                min_row, max_row = ws.min_row, ws.max_row
                min_col, max_col = ws.min_column, ws.max_column
                rows = ws.iter_rows(
                    min_row=min_row, max_row=max_row,
                    min_col=min_col, max_col=max_col
                )

            raw_data: list[list[Any]] = []
            for row in rows:
                row_values = []
                for cell in row:
                    value = cell.value
                    if value is None and not include_empty_cells:
                        continue
                    row_values.append(value)
                raw_data.append(row_values)

            if not raw_data:
                return {
                    "data": [],
                    "sheet_name": actual_sheet_name,
                    "sheet_names": sheet_names,
                    "row_count": 0,
                    "column_count": 0,
                }

            headers: list[str] = []
            data_rows: list[list[Any]] = []

            if has_header and raw_data:
                headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(raw_data[0])]
                data_rows = raw_data[1:]
            else:
                headers = [f"Column_{i}" for i in range(len(raw_data[0]))]
                data_rows = raw_data

            if output_format == "dict":
                output_data: dict[str, list[Any]] = {h: [] for h in headers}
                for row in data_rows:
                    for i, header in enumerate(headers):
                        value = row[i] if i < len(row) else None
                        output_data[header].append(value)
                formatted_data: Any = output_data

            elif output_format == "records":
                formatted_data = []
                for row in data_rows:
                    record = {}
                    for i, header in enumerate(headers):
                        value = row[i] if i < len(row) else None
                        record[header] = value
                    formatted_data.append(record)

            else:  # list
                formatted_data = [headers] + data_rows if has_header else raw_data

            result = {
                "data": formatted_data,
                "headers": headers,
                "sheet_name": actual_sheet_name,
                "sheet_names": sheet_names,
                "row_count": len(data_rows),
                "column_count": len(headers),
                "output_format": output_format,
            }

            if cell_range:
                result["cell_range"] = cell_range

            logger.info(
                "Excel file extraction complete",
                file_path=str(file_path),
                sheet=actual_sheet_name,
                rows=result["row_count"],
                columns=result["column_count"],
            )

            return result

        finally:
            wb.close()
