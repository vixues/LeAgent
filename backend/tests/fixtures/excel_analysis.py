"""Deterministic Excel fixture used by the DeepSeek live integration test.

The workbook contains three sheets that are rich enough to ask
meaningful analysis questions against, but small enough that the model
doesn't need tool-call orchestration just to see the data:

- ``Sales``       — 12 rows, monthly 2024 revenue per region + a total column.
- ``Products``    — 5 rows, product catalog with unit price + category.
- ``Summary``     — a few named metrics (total revenue, best region, worst
  region, top product).

The file lives on disk at
``leagent/backend/tests/fixtures/_cache/excel_sample.xlsx`` so both
``tests/integration/test_deepseek_excel.py`` and
``scripts/run_excel_demo.py`` can point at the same artefact. The cache
is regenerated idempotently whenever the source schema below changes.
"""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pytest


# ---------------------------------------------------------------------------
# Canonical schema (single source of truth for the fixture + expected answers)
# ---------------------------------------------------------------------------


SALES_HEADERS: list[str] = ["Month", "North", "South", "East", "West", "Total"]
SALES_ROWS: list[list[Any]] = [
    ["2024-01", 120_000, 98_000, 110_000, 85_000, 413_000],
    ["2024-02", 135_000, 104_000, 115_000, 92_000, 446_000],
    ["2024-03", 158_000, 112_000, 121_000, 98_000, 489_000],
    ["2024-04", 162_000, 118_000, 128_000, 105_000, 513_000],
    ["2024-05", 175_000, 121_000, 132_000, 110_000, 538_000],
    ["2024-06", 182_000, 125_000, 140_000, 115_000, 562_000],
    ["2024-07", 190_000, 132_000, 145_000, 120_000, 587_000],
    ["2024-08", 195_000, 140_000, 150_000, 125_000, 610_000],
    ["2024-09", 188_000, 138_000, 148_000, 122_000, 596_000],
    ["2024-10", 202_000, 145_000, 155_000, 130_000, 632_000],
    ["2024-11", 215_000, 152_000, 162_000, 138_000, 667_000],
    ["2024-12", 248_000, 175_000, 185_000, 158_000, 766_000],
]

PRODUCTS_HEADERS: list[str] = ["SKU", "Name", "Category", "Unit Price", "In Stock"]
PRODUCTS_ROWS: list[list[Any]] = [
    ["P-001", "Widget Pro",   "Hardware",  49.99,  1200],
    ["P-002", "Gadget Max",   "Hardware", 129.00,   350],
    ["P-003", "Gizmo Lite",   "Hardware",  19.50,  4200],
    ["P-004", "Insight Plus", "Software", 299.00,  9999],
    ["P-005", "Reporter X",   "Software",  79.00,  9999],
]

SUMMARY_ROWS: list[list[Any]] = [
    ["Report Date",    "2024-12-31"],
    ["Total Revenue",  sum(row[-1] for row in SALES_ROWS)],
    ["Best Region",    "North"],      # highest column-sum across SALES_ROWS
    ["Worst Region",   "West"],       # lowest column-sum
    ["Top Product",    "Insight Plus"],  # highest unit price × in-stock > 0
]


@dataclass(frozen=True)
class ExcelFixtureManifest:
    """Canonical shape of the Excel fixture — what the integration test expects."""

    path: Path
    sheets: tuple[str, ...]
    total_revenue: int
    best_region: str
    worst_region: str
    top_product: str
    months: int
    products: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": str(self.path),
            "sheets": list(self.sheets),
            "total_revenue": self.total_revenue,
            "best_region": self.best_region,
            "worst_region": self.worst_region,
            "top_product": self.top_product,
            "months": self.months,
            "products": self.products,
        }


# ---------------------------------------------------------------------------
# Build + cache helpers (importable from scripts, not just pytest)
# ---------------------------------------------------------------------------


def _schema_digest() -> str:
    """Hash the canonical schema so the cache invalidates when content changes."""
    blob = json.dumps(
        {
            "sales": [SALES_HEADERS, SALES_ROWS],
            "products": [PRODUCTS_HEADERS, PRODUCTS_ROWS],
            "summary": SUMMARY_ROWS,
        },
        sort_keys=True,
        default=str,
    ).encode("utf-8")
    return hashlib.sha256(blob).hexdigest()[:12]


def _build_workbook(path: Path) -> None:
    """Write the canonical workbook to ``path`` using openpyxl."""
    import openpyxl

    wb = openpyxl.Workbook()

    sales = wb.active
    sales.title = "Sales"
    sales.append(SALES_HEADERS)
    for row in SALES_ROWS:
        sales.append(row)

    products = wb.create_sheet("Products")
    products.append(PRODUCTS_HEADERS)
    for row in PRODUCTS_ROWS:
        products.append(row)

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    for row in SUMMARY_ROWS:
        summary.append(row)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def ensure_excel_sample(cache_dir: Path | None = None) -> ExcelFixtureManifest:
    """Create ``excel_sample.xlsx`` on disk (idempotent) and return a manifest.

    The file lives at ``<cache_dir>/excel_sample.xlsx`` (default:
    ``tests/fixtures/_cache``) and is regenerated whenever the canonical
    schema digest changes.
    """
    if cache_dir is None:
        cache_dir = Path(__file__).resolve().parent / "_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)

    path = cache_dir / "excel_sample.xlsx"
    digest_file = cache_dir / "excel_sample.digest"
    digest = _schema_digest()

    if not path.exists() or not digest_file.exists() or digest_file.read_text().strip() != digest:
        _build_workbook(path)
        digest_file.write_text(digest, encoding="utf-8")

    return ExcelFixtureManifest(
        path=path,
        sheets=("Sales", "Products", "Summary"),
        total_revenue=int(sum(row[-1] for row in SALES_ROWS)),
        best_region="North",
        worst_region="West",
        top_product="Insight Plus",
        months=len(SALES_ROWS),
        products=len(PRODUCTS_ROWS),
    )


# ---------------------------------------------------------------------------
# Pytest fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(scope="session")
def excel_analysis_manifest(tmp_path_factory: pytest.TempPathFactory) -> ExcelFixtureManifest:
    """Session-scoped Excel fixture.

    Uses the persistent cache under ``tests/fixtures/_cache`` so the demo
    script and the live integration test share the exact same bytes.
    """
    return ensure_excel_sample()


@pytest.fixture(scope="session")
def excel_analysis_path(excel_analysis_manifest: ExcelFixtureManifest) -> Path:
    return excel_analysis_manifest.path
