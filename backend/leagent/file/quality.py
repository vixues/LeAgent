"""Post-register content quality checks for downloadable tool artifacts.

**Ownership:** :class:`~leagent.services.session.manager.SessionManager` is the
only caller that *runs* these checks at promotion time
(``_ensure_quality_on_attachment``). Tool code and QueryEngine must only
*read* ``quality_passed`` / ``quality_error`` from the attachment row.

The helpers below remain for tests and for assessing a path before promotion
when no session is bound yet.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any


@dataclass(frozen=True, slots=True)
class ArtifactQualityResult:
    """Outcome of a content-level artifact quality check."""

    passed: bool
    artifact_type: str
    message: str = ""
    details: dict[str, Any] | None = None


def assess_artifact_quality(
    path: Path | str,
    *,
    content_type: str | None = None,
    filename: str | None = None,
) -> ArtifactQualityResult | None:
    """Return a quality verdict for *path*, or ``None`` if no checker applies."""
    target = Path(path)
    if not target.is_file():
        return ArtifactQualityResult(
            passed=False,
            artifact_type="file",
            message=f"artifact path is not a readable file: {target}",
        )

    name = (filename or target.name).lower()
    mime = (content_type or "").lower()
    suffix = target.suffix.lower()

    if (
        suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}
        or "spreadsheetml" in mime
        or name.endswith(".xlsx")
    ):
        return _assess_xlsx(target)
    if (
        suffix in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
        or mime.startswith("image/")
    ):
        return _assess_image(target)
    if suffix == ".pdf" or mime == "application/pdf":
        return _assess_pdf(target)
    return None


def assess_managed_attachment(attachment: dict[str, Any]) -> ArtifactQualityResult | None:
    """Assess a session attachment dict (prefers ``storage_path``)."""
    storage = (
        attachment.get("storage_path")
        or attachment.get("file_path")
        or attachment.get("source_tool_path")
    )
    if not isinstance(storage, str) or not storage.strip():
        return None
    return assess_artifact_quality(
        storage,
        content_type=attachment.get("content_type")
        if isinstance(attachment.get("content_type"), str)
        else None,
        filename=attachment.get("filename") or attachment.get("name"),
    )


def _assess_xlsx(path: Path) -> ArtifactQualityResult:
    try:
        from openpyxl import load_workbook
    except ImportError:
        # Soft-skip when openpyxl is unavailable — do not block the turn.
        return ArtifactQualityResult(
            passed=True,
            artifact_type="spreadsheet",
            message="openpyxl unavailable; skipped spreadsheet quality check",
        )

    try:
        size = path.stat().st_size
    except OSError as exc:
        return ArtifactQualityResult(
            passed=False,
            artifact_type="spreadsheet",
            message=f"cannot stat spreadsheet: {exc}",
        )
    if size < 64:
        return ArtifactQualityResult(
            passed=False,
            artifact_type="spreadsheet",
            message=f"spreadsheet is suspiciously small ({size} bytes)",
            details={"size": size},
        )

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — surface as quality fail
        return ArtifactQualityResult(
            passed=False,
            artifact_type="spreadsheet",
            message=f"spreadsheet could not be opened: {exc}",
        )

    try:
        data_rows = 0
        header_only_sheets = 0
        sheet_count = 0
        for ws in wb.worksheets:
            sheet_count += 1
            rows_with_values = 0
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    rows_with_values += 1
            if rows_with_values <= 1:
                header_only_sheets += 1
            else:
                data_rows += rows_with_values - 1
    finally:
        wb.close()

    if sheet_count == 0:
        return ArtifactQualityResult(
            passed=False,
            artifact_type="spreadsheet",
            message="spreadsheet has no worksheets",
        )
    if data_rows <= 0:
        return ArtifactQualityResult(
            passed=False,
            artifact_type="spreadsheet",
            message=(
                "spreadsheet appears header-only (no data rows). "
                "Regenerate with real row content, re-register, and cite the "
                "new file_id / download_url."
            ),
            details={
                "sheets": sheet_count,
                "header_only_sheets": header_only_sheets,
                "data_rows": data_rows,
            },
        )
    return ArtifactQualityResult(
        passed=True,
        artifact_type="spreadsheet",
        details={"sheets": sheet_count, "data_rows": data_rows},
    )


def _assess_image(path: Path) -> ArtifactQualityResult:
    try:
        size = path.stat().st_size
    except OSError as exc:
        return ArtifactQualityResult(
            passed=False,
            artifact_type="image",
            message=f"cannot stat image: {exc}",
        )
    if size < 8:
        return ArtifactQualityResult(
            passed=False,
            artifact_type="image",
            message=f"image file is empty or truncated ({size} bytes)",
            details={"size": size},
        )

    try:
        from PIL import Image
    except ImportError:
        # Header sniff without Pillow.
        try:
            head = path.read_bytes()[:16]
        except OSError as exc:
            return ArtifactQualityResult(
                passed=False,
                artifact_type="image",
                message=f"cannot read image: {exc}",
            )
        if not (
            head.startswith(b"\x89PNG")
            or head.startswith(b"\xff\xd8\xff")
            or head[:6] in (b"GIF87a", b"GIF89a")
            or head.startswith(b"RIFF")
        ):
            return ArtifactQualityResult(
                passed=False,
                artifact_type="image",
                message="image file header is not a recognised format",
            )
        return ArtifactQualityResult(passed=True, artifact_type="image")

    try:
        with Image.open(path) as img:
            width, height = img.size
            img.verify()
    except Exception as exc:  # noqa: BLE001
        return ArtifactQualityResult(
            passed=False,
            artifact_type="image",
            message=f"image could not be decoded: {exc}",
        )
    if width <= 0 or height <= 0:
        return ArtifactQualityResult(
            passed=False,
            artifact_type="image",
            message=f"image has invalid dimensions {width}x{height}",
            details={"width": width, "height": height},
        )
    return ArtifactQualityResult(
        passed=True,
        artifact_type="image",
        details={"width": width, "height": height},
    )


def _assess_pdf(path: Path) -> ArtifactQualityResult:
    try:
        size = path.stat().st_size
    except OSError as exc:
        return ArtifactQualityResult(
            passed=False,
            artifact_type="pdf",
            message=f"cannot stat pdf: {exc}",
        )
    if size < 16:
        return ArtifactQualityResult(
            passed=False,
            artifact_type="pdf",
            message=f"pdf file is empty or truncated ({size} bytes)",
            details={"size": size},
        )

    try:
        head = path.read_bytes()[:8]
    except OSError as exc:
        return ArtifactQualityResult(
            passed=False,
            artifact_type="pdf",
            message=f"cannot read pdf: {exc}",
        )
    if not head.startswith(b"%PDF"):
        return ArtifactQualityResult(
            passed=False,
            artifact_type="pdf",
            message="file does not start with a PDF header",
        )

    # Best-effort page count with pypdf / PyPDF2 when available.
    try:
        try:
            from pypdf import PdfReader
        except ImportError:
            from PyPDF2 import PdfReader  # type: ignore[import-not-found]
        reader = PdfReader(str(path))
        pages = len(reader.pages)
        if pages < 1:
            return ArtifactQualityResult(
                passed=False,
                artifact_type="pdf",
                message="pdf has no pages",
                details={"pages": pages},
            )
        return ArtifactQualityResult(
            passed=True,
            artifact_type="pdf",
            details={"pages": pages},
        )
    except Exception:
        # Header OK is enough when no PDF library is installed / parse is soft.
        return ArtifactQualityResult(passed=True, artifact_type="pdf")


def annotate_attachments_quality(
    attachments: list[dict[str, Any]],
) -> tuple[bool, list[str]]:
    """Run quality checks on managed attachment dicts.

    Mutates each attachment in-place with ``quality_passed`` / ``quality_error``
    when a checker applies. Returns ``(all_passed, failure_messages)``.
    """
    failures: list[str] = []
    assessed_any = False
    for att in attachments:
        if not isinstance(att, dict):
            continue
        # Prefer source sandbox path for overwrite checks when still present;
        # otherwise assess the managed storage blob.
        verdict = assess_managed_attachment(att)
        if verdict is None and isinstance(att.get("source_tool_path"), str):
            verdict = assess_artifact_quality(
                att["source_tool_path"],
                content_type=att.get("content_type")
                if isinstance(att.get("content_type"), str)
                else None,
                filename=att.get("filename") or att.get("name"),
            )
        if verdict is None:
            continue
        assessed_any = True
        att["quality_passed"] = verdict.passed
        att["artifact_quality_type"] = verdict.artifact_type
        if verdict.details:
            att["quality_details"] = dict(verdict.details)
        if not verdict.passed:
            msg = verdict.message or "artifact failed quality gate"
            att["quality_error"] = msg
            name = att.get("filename") or att.get("name") or att.get("id") or "file"
            failures.append(f"{name}: {msg}")
    if not assessed_any:
        return True, []
    return (len(failures) == 0), failures


def apply_quality_to_tool_payload(
    data: dict[str, Any],
    *,
    attachments: list[dict[str, Any]] | None = None,
    paths: list[str] | None = None,
) -> bool:
    """Annotate *data* with ``quality_passed`` / ``quality_error``.

    Returns whether all applicable checks passed (True when nothing assessed).
    """
    checked: list[dict[str, Any]] = []
    if attachments:
        checked.extend(a for a in attachments if isinstance(a, dict))
    for raw in paths or []:
        if isinstance(raw, str) and raw.strip():
            checked.append({"storage_path": raw, "filename": Path(raw).name})
    # Also assess produced_files entries still pointing at disk.
    produced = data.get("produced_files")
    if isinstance(produced, list):
        for entry in produced:
            if not isinstance(entry, dict):
                continue
            path = entry.get("source_path") or entry.get("file_path") or entry.get("path")
            if isinstance(path, str) and path.strip():
                checked.append(
                    {
                        "storage_path": path,
                        "filename": entry.get("name") or Path(path).name,
                        "content_type": entry.get("mime") or entry.get("content_type"),
                    }
                )
    for key in ("output_path", "file_path", "path"):
        raw = data.get(key)
        if isinstance(raw, str) and raw.strip():
            checked.append({"storage_path": raw, "filename": Path(raw).name})

    passed, failures = annotate_attachments_quality(checked)
    # Mirror onto produced_files when we annotated matching paths.
    if isinstance(produced, list):
        by_path: dict[str, dict[str, Any]] = {}
        for att in checked:
            sp = att.get("storage_path") or att.get("source_tool_path")
            if isinstance(sp, str):
                try:
                    by_path[str(Path(sp).expanduser().resolve())] = att
                except OSError:
                    by_path[sp] = att
        for entry in produced:
            if not isinstance(entry, dict):
                continue
            for key in ("source_path", "file_path", "path"):
                raw = entry.get(key)
                if not isinstance(raw, str):
                    continue
                try:
                    resolved = str(Path(raw).expanduser().resolve())
                except OSError:
                    resolved = raw
                att = by_path.get(resolved)
                if att is None:
                    continue
                if "quality_passed" in att:
                    entry["quality_passed"] = att["quality_passed"]
                if att.get("quality_error"):
                    entry["quality_error"] = att["quality_error"]
                break

    data["quality_passed"] = passed
    if failures:
        data["quality_error"] = "; ".join(failures)[:1000]
        data["quality_failures"] = failures[:16]
    return passed


def merge_quality_into_tool_data(
    data: dict[str, Any],
    *,
    attachments: list[dict[str, Any]] | None = None,
    paths: list[str] | None = None,
) -> dict[str, Any]:
    """Compatibility wrapper around :func:`apply_quality_to_tool_payload`."""
    apply_quality_to_tool_payload(data, attachments=attachments, paths=paths)
    return data
